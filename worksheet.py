#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Mon Jun 18 22:14:37 2018

@author: ahura
"""

from collections import defaultdict as dd
from openpyxl import load_workbook, worksheet

# The cell describing the meal, e.g. "dinner saturday"
_mealcell = "B2"

# Cell describing the dish, e.g. "Spaghet"
_dishcell = "B3"

# Cell that contains the number of people partaking in the meal
_n_peoplecell = "B5"

# Row number where ingredients start appearing
first_ing_row = 13

ingredient_col = "A"
total_col = "D"
unit_col = "E"
descr_col = "F"

def convert_units(amount, unit):
    if unit == "g":
        return 0.001*amount, "kg"

    return amount, unit

def parse_spreadsheet(filename):
    '''Takes the name of a worksheet file and parses it.
    Returns dict mapping ingredients to
    {'descriptions': [list of descriptions],
    'quantities': {unit: amount}}'''

    # Initialize
    wb = load_workbook(filename=filename, data_only = True)

    data = dd(lambda: {'descriptions': [],
                       'quantities': dd(lambda: 0.0)})
#    data = defaultdict(lambda: defaultdict(lambda: 0.0))
#    details = defaultdict(lambda: [])

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        meal = ws[_mealcell].value
        dish = ws[_dishcell].value
        print("Parsing sheet: %s. Recipe for %s (%s)"
              % (sheet_name, dish, meal))

        for i, row in enumerate(ws.iter_rows()):
            #  Skip rows until we hit the ingredient list
            row_num = i + 1
            if row_num < first_ing_row:
                continue

            # Extract ingredient, amount, and unit.
            col2val = {c.column: c.value for c in row if c.value is not None}
            # Skip empty rows or rows containing name of a subdish
#            if len(col2val) < 2:
#                continue

            # Update quantities
            if ingredient_col not in col2val:
                continue
            ingredient = col2val[ingredient_col]
            # If there's a quantity, add it up
            try:
                amount = col2val[total_col]
                unit_raw = col2val[unit_col] if unit_col in col2val else ""
                amount, unit = convert_units(amount, unit_raw)
                data[ingredient]['quantities'][unit] += amount
            except KeyError:
                pass

            # Update descriptions
            if descr_col not in col2val and ingredient not in data:
                continue  # Skip if no quantity or description
            detail = col2val[descr_col] if descr_col in col2val else "%s (%s)" % (dish, meal)
            data[ingredient]['descriptions'].append(detail)
        #

    return data

if __name__ == '__main__':
    a = parse_spreadsheet("recipes.xlsx")
    print(a)